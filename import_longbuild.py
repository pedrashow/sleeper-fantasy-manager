#!/usr/bin/env python3
"""
Importa rankings da aba 'Long Build PPR Cheatsheet' de um xlsx.
Extrai dados + cor de fundo das células pra mapear tiers.
Gera CSV e/ou insere no SQLite.

Uso:
    python import_longbuild.py caminho/do/arquivo.xlsx
    python import_longbuild.py caminho/do/arquivo.xlsx --db data/fantasy.db
"""

import sys
import os
import csv
import json
import sqlite3
from collections import defaultdict

try:
    from openpyxl import load_workbook
    from openpyxl.styles.colors import Color
except ImportError:
    print("Instale openpyxl: pip install openpyxl")
    sys.exit(1)


SHEET_NAME = "Long Build PPR Cheatsheet"

TIER_COLORS = {
    "beyond":        {"name": "Beyond Category",                          "tier": 1},
    "gold":          {"name": "Gold Standard: Elite",                     "tier": 2},
    "blue_chip":     {"name": "Blue Chip: Top Starter",                   "tier": 3},
    "reliable":      {"name": "Reliable/Emerging: Capable of Elite #s",   "tier": 4},
    "immediate":     {"name": "Immediate Starter Talent",                 "tier": 5},
    "contributor":   {"name": "Contributors/Low-End Starters w/ Upside",  "tier": 6},
    "reserve":       {"name": "Reserves Capable of Production",           "tier": 7},
    "developmental": {"name": "Developmental, Bubble, Role Players",      "tier": 8},
}


def get_cell_rgb(cell):
    fill = cell.fill
    if fill is None or fill.fgColor is None:
        return None
    color = fill.fgColor
    if color.type == "rgb" and color.rgb and color.rgb != "00000000":
        rgb = str(color.rgb)
        if len(rgb) == 8:
            rgb = rgb[2:]
        if len(rgb) == 6:
            try:
                return tuple(int(rgb[i:i+2], 16) for i in (0, 2, 4))
            except ValueError:
                return None
    if color.type == "theme":
        return ("theme", color.theme, color.tint if color.tint else 0)
    if color.type == "indexed":
        return ("indexed", color.indexed)
    return None


def detect_position_blocks(ws, header_row=2):
    blocks = []
    row = list(ws.iter_rows(min_row=header_row, max_row=header_row))[0]
    for cell in row:
        val = str(cell.value or "").strip().lower()
        if "quarterback" in val:
            blocks.append({"pos": "QB", "col_start": cell.column})
        elif "running back" in val:
            blocks.append({"pos": "RB", "col_start": cell.column})
        elif "wide receiver" in val:
            blocks.append({"pos": "WR", "col_start": cell.column})
        elif "tight end" in val:
            blocks.append({"pos": "TE", "col_start": cell.column})
    return blocks


def find_header_row(ws):
    for row_num in range(1, 6):
        for cell in ws[row_num]:
            val = str(cell.value or "").strip().lower()
            if "quarterback" in val or "running back" in val:
                return row_num
    return 2


NAME_FIXES = {
    "Isiah Pachecho": "Isiah Pacheco",
    "Tyler Allegier": "Tyler Allgeier",
    "James Connor": "James Conner",
    "Jauan Jennngs": "Jauan Jennings",
    "Aiden O'Connell": "Aidan O'Connell",
    "Eli Eidenreich": "Eli Heidenreich",
    "Xavier Guilllory": "Xavier Guillory",
    "Drew Olgetree": "Drew Ogletree",
    "Lan Larisson": "Lan Larison",
    "Kene Ngwangwu": "Kene Nwangwu",
}


def parse_name_team(text):
    if not text:
        return None, None
    text = str(text).strip()
    parts = text.rsplit(" - ", 1)
    if len(parts) == 2:
        name, team = parts[0].strip(), parts[1].strip()
    else:
        parts = text.rsplit("-", 1)
        if len(parts) == 2:
            name, team = parts[0].strip(), parts[1].strip()
        else:
            name, team = text, ""

    if "," in name:
        name_parts = name.split(",", 1)
        name = f"{name_parts[1].strip()} {name_parts[0].strip()}"

    name = NAME_FIXES.get(name, name)

    team_map = {"UFA": "FA", "49ers": "SF", "JAC": "JAX"}
    team = team_map.get(team, team)

    return name, team


def extract_block(ws, block, header_row, all_colors):
    pos = block["pos"]
    rank_col = block["col_start"] - 1
    name_col = block["col_start"]
    age_col = block["col_start"] + 1
    notes_col = block["col_start"] + 2

    players = []
    pos_rank = 0

    for row_num in range(header_row + 1, ws.max_row + 1):
        name_cell = ws.cell(row=row_num, column=name_col)
        rank_cell = ws.cell(row=row_num, column=rank_col)
        age_cell = ws.cell(row=row_num, column=age_col)
        notes_cell = ws.cell(row=row_num, column=notes_col)

        name_val = name_cell.value
        if name_val is None or str(name_val).strip() == "":
            continue

        full_name, team = parse_name_team(name_val)
        if not full_name:
            continue

        overall_rank = rank_cell.value
        try:
            overall_rank = int(overall_rank)
        except (TypeError, ValueError):
            overall_rank = None

        age = age_cell.value
        try:
            age = int(age)
        except (TypeError, ValueError):
            age = None

        notes = str(notes_cell.value or "").strip()

        rgb = get_cell_rgb(name_cell)
        if rgb is None:
            rgb = get_cell_rgb(rank_cell)

        pos_rank += 1

        player = {
            "position": pos,
            "pos_rank": pos_rank,
            "overall_rank": overall_rank,
            "full_name": full_name,
            "team": team,
            "age": age,
            "notes": notes,
            "cell_color": rgb,
        }
        players.append(player)

        if rgb and not isinstance(rgb, tuple):
            pass
        elif rgb:
            all_colors.append(rgb)

    return players


def auto_assign_tiers(players):
    color_groups = defaultdict(list)
    for i, p in enumerate(players):
        c = p.get("cell_color")
        key = str(c) if c else "none"
        color_groups[key].append(i)

    ordered_groups = []
    seen = set()
    for i, p in enumerate(players):
        key = str(p.get("cell_color"))
        if key not in seen:
            seen.add(key)
            ordered_groups.append(key)

    tier_num = 0
    for group_key in ordered_groups:
        tier_num += 1
        for idx in color_groups[group_key]:
            players[idx]["tier"] = tier_num

    return players


def print_color_summary(all_players):
    print("\n=== CORES DETECTADAS (por tier) ===")
    tier_colors = defaultdict(set)
    tier_players = defaultdict(list)
    for p in all_players:
        t = p.get("tier", "?")
        c = p.get("cell_color")
        tier_colors[t].add(str(c))
        tier_players[t].append(p["full_name"])

    for tier in sorted(tier_colors.keys()):
        colors = tier_colors[tier]
        players = tier_players[tier]
        sample = ", ".join(players[:3])
        if len(players) > 3:
            sample += f" +{len(players)-3}"
        print(f"  Tier {tier}: {len(players)} jogadores | Cores: {colors}")
        print(f"    Ex: {sample}")


def save_csv(all_players, output_path):
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "position", "pos_rank", "overall_rank", "full_name",
            "team", "age", "notes", "tier"
        ])
        writer.writeheader()
        for p in all_players:
            row = {k: v for k, v in p.items() if k != "cell_color"}
            writer.writerow(row)
    print(f"\nCSV salvo: {output_path}")


import re
import unicodedata

def normalize_name(name):
    if not name:
        return ""
    name = unicodedata.normalize("NFKD", name)
    name = name.lower().strip()
    name = name.replace(".", "").replace("'", "").replace("'", "").replace("-", " ")
    name = re.sub(r"\s+", " ", name)
    name = re.sub(r"\b(jr|sr|ii|iii|iv)\b", "", name).strip()
    return name


def match_player_id(conn, full_name, team, position):
    row = conn.execute(
        "SELECT player_id FROM players WHERE full_name = ? AND position = ?",
        (full_name, position)
    ).fetchone()
    if row:
        return row[0]

    norm = normalize_name(full_name)
    parts = norm.split()
    if len(parts) < 2:
        return None

    first = parts[0]
    last = parts[-1]

    rows = conn.execute("""
        SELECT player_id, full_name, team FROM players
        WHERE position = ?
        AND LOWER(full_name) LIKE ?
        AND LOWER(full_name) LIKE ?
    """, (position, f"%{first}%", f"%{last}%")).fetchall()

    if len(rows) == 1:
        return rows[0][0]

    if len(rows) > 1 and team and team != "FA":
        for r in rows:
            if r[2] and r[2].upper() == team.upper():
                return r[0]

    if len(rows) > 1:
        for r in rows:
            r_norm = normalize_name(r[1])
            if r_norm == norm:
                return r[0]

    if not rows:
        rows = conn.execute("""
            SELECT player_id, full_name, team FROM players
            WHERE position = ? AND LOWER(full_name) LIKE ?
        """, (position, f"%{last}%")).fetchall()

        if len(rows) == 1:
            return rows[0][0]
        if len(rows) > 1 and team and team != "FA":
            for r in rows:
                if r[2] and r[2].upper() == team.upper():
                    return r[0]

    return None


def save_to_db(all_players, db_path, table_name="longbuild_rankings"):
    conn = sqlite3.connect(db_path)
    conn.execute(f"""
        CREATE TABLE IF NOT EXISTS {table_name} (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            player_id TEXT,
            position TEXT,
            pos_rank INTEGER,
            overall_rank INTEGER,
            full_name TEXT,
            team TEXT,
            age INTEGER,
            notes TEXT,
            tier INTEGER,
            tier_name TEXT,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.execute(f"DELETE FROM {table_name}")

    matched = 0
    unmatched = []

    for p in all_players:
        tier = p.get("tier", 99)
        tier_name = ""
        for k, v in TIER_COLORS.items():
            if v["tier"] == tier:
                tier_name = v["name"]
                break

        player_id = match_player_id(conn, p["full_name"], p["team"], p["position"])
        if player_id:
            matched += 1
        else:
            unmatched.append(f"{p['full_name']} ({p['position']} - {p['team']})")

        conn.execute(f"""
            INSERT INTO {table_name}
            (player_id, position, pos_rank, overall_rank, full_name, team, age, notes, tier, tier_name)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            player_id, p["position"], p["pos_rank"], p.get("overall_rank"),
            p["full_name"], p["team"], p.get("age"),
            p.get("notes", ""), tier, tier_name
        ))

    conn.commit()
    count = conn.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
    print(f"\n{count} jogadores inseridos em {table_name}")
    print(f"  ✅ {matched} com player_id (matched)")
    if unmatched:
        print(f"  ❌ {len(unmatched)} sem match:")
        for u in unmatched:
            print(f"     {u}")
    conn.close()


SHEETS = {
    "longbuild": {"sheet": "Long Build PPR Cheatsheet", "table": "longbuild_rankings"},
    "winnow":    {"sheet": "Win-Now PPR Cheatsheet",    "table": "winnow_rankings"},
}


def import_sheet(xlsx_path, sheet_key="all", db_path="data/fantasy.db"):
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"File not found: {xlsx_path}")

    print(f"Opening {xlsx_path}...")
    wb = load_workbook(xlsx_path, data_only=True)

    if sheet_key == "all":
        sheets_to_import = list(SHEETS.keys())
    elif sheet_key in SHEETS:
        sheets_to_import = [sheet_key]
    else:
        raise ValueError(f"Unknown sheet: {sheet_key}. Options: {', '.join(SHEETS.keys())}, all")

    for key in sheets_to_import:
        sheet_name = SHEETS[key]["sheet"]
        table_name = SHEETS[key]["table"]

        print(f"\n{'='*50}")
        print(f"Importing: {sheet_name} → {table_name}")
        print(f"{'='*50}")

        if sheet_name not in wb.sheetnames:
            print(f"  Sheet '{sheet_name}' not found. Skipping.")
            continue

        ws = wb[sheet_name]
        header_row = find_header_row(ws)
        blocks = detect_position_blocks(ws, header_row)
        if not blocks:
            print("  No position blocks found. Skipping.")
            continue

        all_colors = []
        all_players = []

        for block in blocks:
            players = extract_block(ws, block, header_row, all_colors)
            print(f"    {block['pos']}: {len(players)} players")
            all_players.extend(players)

        for pos in ["QB", "RB", "WR", "TE"]:
            pos_players = [p for p in all_players if p["position"] == pos]
            pos_players = auto_assign_tiers(pos_players)
            for pp in pos_players:
                for ap in all_players:
                    if ap["full_name"] == pp["full_name"] and ap["position"] == pp["position"]:
                        ap["tier"] = pp["tier"]

        print(f"  Total: {len(all_players)} players")
        save_to_db(all_players, db_path, table_name)


def main():
    if len(sys.argv) < 2:
        print("Uso:")
        print("  python import_longbuild.py arquivo.xlsx                     (importa Long Build)")
        print("  python import_longbuild.py arquivo.xlsx --sheet winnow      (importa Win-Now)")
        print("  python import_longbuild.py arquivo.xlsx --sheet all         (importa ambas)")
        print("  python import_longbuild.py arquivo.xlsx --db data/fantasy.db")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    db_path = None
    sheet_key = "longbuild"

    if "--db" in sys.argv:
        db_idx = sys.argv.index("--db")
        if db_idx + 1 < len(sys.argv):
            db_path = sys.argv[db_idx + 1]

    if "--sheet" in sys.argv:
        s_idx = sys.argv.index("--sheet")
        if s_idx + 1 < len(sys.argv):
            sheet_key = sys.argv[s_idx + 1]

    if not os.path.exists(xlsx_path):
        print(f"Arquivo não encontrado: {xlsx_path}")
        sys.exit(1)

    print(f"Abrindo {xlsx_path}...")
    wb = load_workbook(xlsx_path, data_only=True)

    if sheet_key == "all":
        sheets_to_import = list(SHEETS.keys())
    elif sheet_key in SHEETS:
        sheets_to_import = [sheet_key]
    else:
        print(f"Sheet desconhecida: {sheet_key}")
        print(f"Opções: {', '.join(SHEETS.keys())}, all")
        sys.exit(1)

    for key in sheets_to_import:
        sheet_name = SHEETS[key]["sheet"]
        table_name = SHEETS[key]["table"]

        print(f"\n{'='*50}")
        print(f"Importando: {sheet_name} → {table_name}")
        print(f"{'='*50}")

        if sheet_name not in wb.sheetnames:
            print(f"  Aba '{sheet_name}' não encontrada. Pulando.")
            print(f"  Abas disponíveis: {wb.sheetnames}")
            continue

        ws = wb[sheet_name]
        print(f"  Aba encontrada ({ws.max_row} linhas x {ws.max_column} colunas)")

        header_row = find_header_row(ws)
        print(f"  Header na linha {header_row}")

        blocks = detect_position_blocks(ws, header_row)
        if not blocks:
            print("  Nenhum bloco de posição encontrado. Pulando.")
            continue

        print(f"  Blocos: {[b['pos'] for b in blocks]}")

        all_colors = []
        all_players = []

        for block in blocks:
            players = extract_block(ws, block, header_row, all_colors)
            print(f"    {block['pos']}: {len(players)} jogadores")
            all_players.extend(players)

        for pos in ["QB", "RB", "WR", "TE"]:
            pos_players = [p for p in all_players if p["position"] == pos]
            pos_players = auto_assign_tiers(pos_players)
            for pp in pos_players:
                for ap in all_players:
                    if ap["full_name"] == pp["full_name"] and ap["position"] == pp["position"]:
                        ap["tier"] = pp["tier"]

        print(f"\n  Total: {len(all_players)} jogadores")
        print_color_summary(all_players)

        csv_path = xlsx_path.rsplit(".", 1)[0] + f"_{key}.csv"
        save_csv(all_players, csv_path)

        if db_path:
            save_to_db(all_players, db_path, table_name)

        print(f"\n  PREVIEW {key.upper()}:")
        for pos in ["QB", "RB", "WR", "TE"]:
            pos_players = [p for p in all_players if p["position"] == pos]
            print(f"  {pos} ({len(pos_players)}):")
            for p in pos_players[:3]:
                print(f"    {p['pos_rank']:>3}. {p['full_name']:<25} {p['team']:<4} T{p.get('tier','?')} | #{p.get('overall_rank','?')}")
            if len(pos_players) > 3:
                print(f"    ... +{len(pos_players)-3} mais")


if __name__ == "__main__":
    main()
